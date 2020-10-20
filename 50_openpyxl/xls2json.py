from openpyxl import load_workbook
import json


class Data:
    def __init__(self, order_data, region, rep, item, units, unit_cost, total):
        self.orderData = order_data
        self.region = region
        self.rep = rep
        self.item = item
        self.units = units
        self.unitsCost = unit_cost
        self.total = total

    def __repr__(self):
        return f'{self.__dict__}'

    def get_data(self):
        return {
            'order_data': self.orderData,
            'region': self.region,
            'rep': self.rep,
            'item': self.item,
            'units': self.units,
        }


class Sample:
    def __init__(self, file_name, sheet_name='Sheet1'):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.datas = []
        self.read_xls()

    def __repr__(self):
        return f'File name:{self.file_name}, with a list of {len(self.datas)} datas'

    def read_xls(self):
        ws = load_workbook(self.file_name)[self.sheet_name]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            column_value = []
            for cell in row:
                column_value.append(cell.value)
            self.datas.append(
                (
                    Data(
                        column_value[0],
                        column_value[1],
                        column_value[2],
                        column_value[3],
                        column_value[4],
                        column_value[5],
                        column_value[6],
                    )
                )
            )

    def get_datas(self):
        return self.datas


class BestSalesItems:
    def __init__(self, datas):
        self.catalog = {}
        self.datas2itemrep(datas)

    def datas2itemrep(self, datas):
        for data in datas:
            infos = data.get_data()
            if infos['item'] not in self.catalog.keys():
                self.catalog[infos['item']] = {infos['rep']: infos['units']}
            elif infos['rep'] not in self.catalog[infos['item']].keys():
                self.catalog[infos['item']][infos['rep']] = infos['units']
            else:
                value = (
                    self.catalog[infos['item']][infos['rep']] + infos['units']
                )
                self.catalog[infos['item']][infos['rep']] = value

    def catalog2json(self):
        result = {'catalog': self.catalog}
        return json.dumps(result, indent=2)


teste = Sample('SampleData.xlsx', 'SalesOrders')
datas = teste.get_datas()
sales = BestSalesItems(datas)
new_struc = sales.catalog2json()
print(new_struc)
