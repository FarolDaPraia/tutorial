import openpyxl

wb = openpyxl.load_workbook('SampleData.xlsx')
print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x0000024A8D8EB9A0>

type(wb)  # <class 'openpyxl.workbook.workbook.Workbook'>

wb.get_sheet_names()  # class method
print(wb.sheetnames)  # list class with the names of all worksheets
for sheet in wb:  # loop through worksheets
    print(sheet.title)

salesOrders = wb[
    'SalesOrders'
]  # type(salesOrders) <class 'openpyxl.worksheet.worksheet.Worksheet'>

c = salesOrders['A4']  # access cell directly as keys <Cell 'SalesOrders'.A4>
d = salesOrders.cell(
    row=4, column=2, value=10
)  # provides access to cells using row and column notation <Cell 'SalesOrders'.B4>
cell_range = salesOrders[
    'A1':'C2'
]  # Ranges of cells can be accessed using slicing
